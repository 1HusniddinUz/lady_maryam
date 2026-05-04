"""
Excel va grafik eksport xizmati
"""

import io
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")  # headless rejim
import matplotlib.pyplot as plt
import matplotlib.dates as mdates


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header_row(ws, row_num: int, columns: int):
    for col in range(1, columns + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)


def export_sales_report_excel(
    summary: dict,
    top_products_list: List[dict],
    top_customers_list: List[dict],
    daily_data: List[dict],
    period_name: str,
) -> bytes:
    """Sotuv hisobotini Excel formatida"""
    wb = Workbook()

    # ── 1-list: Asosiy hisobot ──
    ws = wb.active
    ws.title = "Asosiy hisobot"
    ws["A1"] = f"Hisobot: {period_name}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Yaratildi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")

    # Asosiy ko'rsatkichlar jadvali
    rows = [
        ("Ko'rsatkich", "Qiymat"),
        ("Buyurtmalar soni", summary["orders_count"]),
        ("Sotilgan mahsulotlar (dona)", summary["items_count"]),
        ("Jami tushum", summary["revenue"]),
        ("Tan narxi", summary["cost"]),
        ("Yalpi foyda", summary["gross_profit"]),
        ("Yalpi marja %", f"{summary['margin_pct']:.1f}%"),
        ("Soliq", summary["tax"]),
        ("Soliqdan keyin foyda", summary["net_after_tax"]),
        ("Operatsion xarajatlar", summary["expenses"]),
        ("YAKUNIY SOF FOYDA", summary["final_net"]),
        ("Sof foyda %", f"{summary['net_pct']:.1f}%"),
    ]
    start_row = 4
    for i, (label, value) in enumerate(rows):
        ws.cell(row=start_row + i, column=1, value=label)
        ws.cell(row=start_row + i, column=2, value=value)
        ws.cell(row=start_row + i, column=1).border = THIN_BORDER
        ws.cell(row=start_row + i, column=2).border = THIN_BORDER
        if i == 0:
            _style_header_row(ws, start_row, 2)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            ws.cell(row=start_row + i, column=2).number_format = '#,##0 "so\'m"'

    # Yakuniy sof foyda qatorini ajratib ko'rsatish
    final_row = start_row + 10
    for col in (1, 2):
        c = ws.cell(row=final_row, column=col)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    _autosize(ws)

    # ── 2-list: TOP mahsulotlar ──
    ws2 = wb.create_sheet("TOP Mahsulotlar")
    ws2["A1"] = "TOP-10 sotilgan mahsulotlar"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:E1")

    headers = ["№", "Mahsulot", "Sotilgan (dona)", "Tushum", "Foyda"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=3, column=i, value=h)
    _style_header_row(ws2, 3, len(headers))

    for idx, p in enumerate(top_products_list, 1):
        ws2.cell(row=3 + idx, column=1, value=idx)
        ws2.cell(row=3 + idx, column=2, value=p["name"])
        ws2.cell(row=3 + idx, column=3, value=p["quantity"])
        ws2.cell(row=3 + idx, column=4, value=p["revenue"]).number_format = '#,##0 "so\'m"'
        ws2.cell(row=3 + idx, column=5, value=p["profit"]).number_format = '#,##0 "so\'m"'
        for col in range(1, 6):
            ws2.cell(row=3 + idx, column=col).border = THIN_BORDER
    _autosize(ws2)

    # ── 3-list: TOP mijozlar ──
    ws3 = wb.create_sheet("TOP Mijozlar")
    ws3["A1"] = "TOP-10 mijozlar"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.merge_cells("A1:E1")
    headers = ["№", "Ism", "Telefon", "Buyurtmalar", "Jami xarid"]
    for i, h in enumerate(headers, 1):
        ws3.cell(row=3, column=i, value=h)
    _style_header_row(ws3, 3, len(headers))
    for idx, c in enumerate(top_customers_list, 1):
        ws3.cell(row=3 + idx, column=1, value=idx)
        ws3.cell(row=3 + idx, column=2, value=c["name"])
        ws3.cell(row=3 + idx, column=3, value=c.get("phone") or "—")
        ws3.cell(row=3 + idx, column=4, value=c["orders"])
        ws3.cell(row=3 + idx, column=5, value=c["spent"]).number_format = '#,##0 "so\'m"'
        for col in range(1, 6):
            ws3.cell(row=3 + idx, column=col).border = THIN_BORDER
    _autosize(ws3)

    # ── 4-list: Kunlik dinamika ──
    ws4 = wb.create_sheet("Kunlik dinamika")
    ws4["A1"] = "Kunlik sotuv dinamikasi"
    ws4["A1"].font = Font(bold=True, size=14)
    ws4.merge_cells("A1:D1")
    headers = ["Sana", "Buyurtmalar", "Tushum", "Foyda"]
    for i, h in enumerate(headers, 1):
        ws4.cell(row=3, column=i, value=h)
    _style_header_row(ws4, 3, len(headers))
    for idx, d in enumerate(daily_data, 1):
        ws4.cell(row=3 + idx, column=1, value=d["date"])
        ws4.cell(row=3 + idx, column=2, value=d["orders"])
        ws4.cell(row=3 + idx, column=3, value=d["revenue"]).number_format = '#,##0 "so\'m"'
        ws4.cell(row=3 + idx, column=4, value=d["profit"]).number_format = '#,##0 "so\'m"'
        for col in range(1, 5):
            ws4.cell(row=3 + idx, column=col).border = THIN_BORDER
    _autosize(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_products_excel(products: List) -> bytes:
    """Mahsulotlar ro'yxatini Excel ga"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"
    ws["A1"] = "Mahsulotlar ro'yxati"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    headers = ["№", "Nomi", "Tan narxi", "Sotuv narxi", "Marja", "Qoldiq", "Birlik"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    for idx, p in enumerate(products, 1):
        ws.cell(row=3 + idx, column=1, value=idx)
        ws.cell(row=3 + idx, column=2, value=p.name)
        ws.cell(row=3 + idx, column=3, value=p.cost_price).number_format = '#,##0 "so\'m"'
        ws.cell(row=3 + idx, column=4, value=p.sale_price).number_format = '#,##0 "so\'m"'
        ws.cell(row=3 + idx, column=5, value=p.margin).number_format = '#,##0 "so\'m"'
        ws.cell(row=3 + idx, column=6, value=p.stock_quantity)
        ws.cell(row=3 + idx, column=7, value=p.unit)
        for col in range(1, 8):
            ws.cell(row=3 + idx, column=col).border = THIN_BORDER

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_sales_chart(daily_data: List[dict], period_name: str) -> bytes:
    """Kunlik sotuv grafigi (PNG)"""
    if not daily_data:
        # Bo'sh holat uchun placeholder grafik
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.text(0.5, 0.5, "Ma'lumot mavjud emas", ha="center", va="center", fontsize=16, color="gray")
        ax.set_axis_off()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight")
        plt.close(fig)
        return buf.getvalue()

    dates = [datetime.fromisoformat(d["date"]) for d in daily_data]
    revenues = [d["revenue"] / 1000 for d in daily_data]   # ming so'mda
    profits = [d["profit"] / 1000 for d in daily_data]

    fig, ax = plt.subplots(figsize=(11, 6))
    ax.plot(dates, revenues, marker="o", linewidth=2, label="Tushum", color="#4472C4")
    ax.plot(dates, profits, marker="s", linewidth=2, label="Sof foyda", color="#70AD47")
    ax.fill_between(dates, revenues, alpha=0.1, color="#4472C4")
    ax.fill_between(dates, profits, alpha=0.1, color="#70AD47")

    ax.set_title(f"Sotuv dinamikasi — {period_name}", fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Sana", fontsize=11)
    ax.set_ylabel("Summa (ming so'm)", fontsize=11)
    ax.legend(loc="upper left", frameon=True)
    ax.grid(True, alpha=0.3, linestyle="--")

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
    fig.autofmt_xdate()

    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def make_top_products_chart(top_products_list: List[dict]) -> bytes:
    """TOP mahsulotlar grafigi"""
    if not top_products_list:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.text(0.5, 0.5, "Ma'lumot yo'q", ha="center", va="center", fontsize=16, color="gray")
        ax.set_axis_off()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight")
        plt.close(fig)
        return buf.getvalue()

    names = [p["name"][:25] for p in top_products_list[:10]]
    qtys = [p["quantity"] for p in top_products_list[:10]]

    fig, ax = plt.subplots(figsize=(10, 6))
    bars = ax.barh(names, qtys, color="#4472C4")
    ax.set_xlabel("Sotilgan miqdor", fontsize=11)
    ax.set_title("TOP-10 sotilgan mahsulotlar", fontsize=14, fontweight="bold", pad=15)
    ax.invert_yaxis()

    for bar in bars:
        w = bar.get_width()
        ax.text(w + max(qtys) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{w:.0f}", va="center", fontsize=10)

    ax.grid(True, alpha=0.3, axis="x", linestyle="--")
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()
